import openpyxl
from datetime import datetime
from .models import Employee, Attendance

def parse_datetime(value):
    if isinstance(value,datetime):
        return value
    value = str(value).strip()
    formats = [
        '%d/%m/%Y %H:%M:%S',
        '%d/%m/%Y %H:%M:%S %p',
        '%d/%m/%Y  %H:%M:%S %p'
    ]
    for fmt in formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    print(f"Could not parse date: {value}")
    return None

def load_excel_data():
    wb = openpyxl.load_workbook('project.xlsx')

    sheet2 = wb['Input reference 2']
    for row in sheet2.iter_rows(min_row = 2, values_only=True):
        employee_id,name,company = row
        if employee_id:
            Employee.objects.get_or_create(
                employee_id = employee_id,
                defaults = {
                    'name' : name,
                    'company' : company
                }
            )
    sheet1 = wb['Input reference 1']
    for row in sheet1.iter_rows(min_row = 2, values_only=True):
        employee_id,start_date,end_date = row[0],row[1],row[2]
        if employee_id and start_date and end_date:
            try:
                employee = Employee.objects.filter(employee_id = employee_id).first()
                if not employee:
                    print("Employee not found")
                    continue

                start_dt = parse_datetime(start_date)
                end_dt = parse_datetime(end_date)

                if not start_dt or not end_dt:
                    print("Start Date and End Date not found")
                    continue

                #Calculating working hours
                diff = end_dt - start_dt
                working_hours = round(diff.total_seconds()/3600,2)

                #Attendance saving
                Attendance.objects.get_or_create(
                employee = employee,
                date = start_dt.date(),
                defaults= {
                    'start_time' : start_dt.time(),
                    'end_time' : end_dt.time(),
                    'working_hours' : working_hours
                }
                )
            except Exception as e:
                print(e)
    print("Excel data loaded successfully")